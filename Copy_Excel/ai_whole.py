import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, PatternFill, Alignment


def copy_excel_with_format(source_file, target_file):
    """
    将源Excel表（单工作表）完整复制到目标Excel表，保留所有格式

    参数:
        source_file (str): 源文件路径（如："美团需求表.xlsx"）
        target_file (str): 目标文件路径（如："美团开放岗位表.xlsx"）
    """
    try:
        # === 1. 文件检查 ===
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"源文件不存在: {source_file}")
        if not os.path.exists(target_file):
            raise FileNotFoundError(f"目标文件不存在: {target_file}")

        # === 2. 加载文件 ===
        source_wb = load_workbook(source_file)
        target_wb = load_workbook(target_file)

        # 获取活动工作表（适用于单工作表文件）
        source_ws = source_wb.active
        target_ws = target_wb.active

        # === 3. 清除目标表原有内容（保留工作表）===
        target_ws.delete_rows(1, target_ws.max_row)  # 删除所有行
        target_ws.delete_cols(1, target_ws.max_column)  # 删除所有列

        # === 4. 复制数据 ===
        for row in source_ws.iter_rows():
            for cell in row:
                # 创建目标单元格
                new_cell = target_ws[cell.coordinate]

                # 复制值
                new_cell.value = cell.value

                # === 5. 复制格式 ===
                if cell.has_style:
                    # 字体
                    new_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )

                    # 边框
                    new_cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom
                    )

                    # 背景色
                    if cell.fill.patternType is not None:
                        new_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        )

                    # 对齐方式
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text
                    )

                # 复制数字格式（如日期、货币等）
                if hasattr(cell, 'number_format'):
                    new_cell.number_format = cell.number_format

        # === 6. 复制列宽 ===
        for col in range(1, source_ws.max_column + 1):
            col_letter = get_column_letter(col)
            target_ws.column_dimensions[col_letter].width = \
                source_ws.column_dimensions[col_letter].width

        # === 7. 复制行高 ===
        for row in range(1, source_ws.max_row + 1):
            if row in source_ws.row_dimensions:
                target_ws.row_dimensions[row].height = \
                    source_ws.row_dimensions[row].height

        # === 8. 保存文件 ===
        target_wb.save(target_file)
        print(f"✅ 成功复制: {source_file} -> {target_file}")

    except Exception as e:
        print(f"❌ 发生错误: {str(e)}")
        raise


# 使用示例
if __name__ == "__main__":
    # 请修改为您的实际文件路径
    copy_excel_with_format(
        source_file="美团需求.xlsx",
        target_file="美团开放岗位.xlsx"
    )