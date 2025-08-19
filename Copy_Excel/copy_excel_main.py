from openpyxl import load_workbook, Workbook
import copy
import os

source_file = "美团需求.xlsx"
target_file = "美团开放岗位.xlsx"

if os.path.exists(source_file) and os.path.exists(target_file):
    # 复制到内存中
    workbook_source = load_workbook(source_file)
    workbook_target = load_workbook(target_file)
    # 获取当前活动的工作表（默认打开的那个）
    sheet_source = workbook_source.active
    sheet_target = workbook_target.active

    # 清空目标表所有内容
    sheet_target.delete_rows(1, sheet_target.max_row)

    # row_index记录行号，row_dimension记录行的高度
    for row_index, row_dimension in sheet_source.row_dimensions.items():
        sheet_target.row_dimensions[row_index].height = row_dimension.height

    # 记录行宽
    for col_index, col_dimension in sheet_source.column_dimensions.items():
        sheet_target.column_dimensions[col_index].width = col_dimension.width

    # 复制单元格
    for row in sheet_source.iter_rows():
        for cell in row:
            new_cell = sheet_target.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy.copy(cell.font)  # 字体
                new_cell.border = copy.copy(cell.border)  # 边框
                new_cell.fill = copy.copy(cell.fill)  # 单元格
                new_cell.number_format = copy.copy(cell.number_format)  # 数字格式
                new_cell.protection = copy.copy(cell.protection)  # 单元格保护格式
                new_cell.alignment = copy.copy(cell.alignment)  # 文字对齐方式

    # 从内存中存到硬盘
    workbook_target.save(target_file)
    print("数据复制完成！")

else:
    print("源文件或者目标文件不存在，请检查路径")


